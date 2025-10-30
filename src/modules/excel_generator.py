"""
Excel File Generator Module
Creates Excel spreadsheets from PDF content
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Optional
import os
import re


class ExcelGenerator:
    """Generate Excel spreadsheets from content"""
    
    def __init__(self, config: dict = None):
        """
        Initialize Excel generator
        
        Args:
            config: Configuration dictionary
        """
        self.config = config or self.get_default_config()
        self.wb = None
        self.ws = None
        
    def get_default_config(self) -> dict:
        """Get default configuration"""
        return {
            'title_font_size': 16,
            'header_font_size': 12,
            'content_font_size': 11,
            'title_color': '003366',
            'header_color': 'CCDDEE',
            'auto_width': True
        }
    
    def create_workbook(self, title: str = "Data") -> Workbook:
        """
        Create a new workbook
        
        Args:
            title: Workbook title
            
        Returns:
            Workbook object
        """
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = title[:31]  # Excel sheet name limit
        return self.wb
    
    def add_title(self, title: str, row: int = 1):
        """
        Add title to worksheet
        
        Args:
            title: Title text
            row: Row number
        """
        self.ws[f'A{row}'] = title
        
        # Format title
        title_cell = self.ws[f'A{row}']
        title_cell.font = Font(
            size=self.config.get('title_font_size', 16),
            bold=True,
            color=self.config.get('title_color', '003366')
        )
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge cells for title
        self.ws.merge_cells(f'A{row}:E{row}')
    
    def add_header_row(self, headers: List[str], row: int = 3):
        """
        Add header row
        
        Args:
            headers: List of header texts
            row: Row number
        """
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col_num)
            cell.value = header
            
            # Format header
            cell.font = Font(
                size=self.config.get('header_font_size', 12),
                bold=True
            )
            cell.fill = PatternFill(
                start_color=self.config.get('header_color', 'CCDDEE'),
                end_color=self.config.get('header_color', 'CCDDEE'),
                fill_type='solid'
            )
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    def add_data_row(self, data: List[str], row: int):
        """
        Add data row
        
        Args:
            data: List of cell values
            row: Row number
        """
        for col_num, value in enumerate(data, 1):
            cell = self.ws.cell(row=row, column=col_num)
            cell.value = value
            
            # Format cell
            cell.font = Font(size=self.config.get('content_font_size', 11))
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    def auto_adjust_column_width(self):
        """Auto-adjust column widths"""
        if not self.config.get('auto_width', True):
            return
        
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            self.ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_from_chunks(self, text_chunks: List[str], title: str,
                            output_path: str = None) -> str:
        """
        Generate Excel file from text chunks
        
        Args:
            text_chunks: List of text chunks
            title: Workbook title
            output_path: Output file path
            
        Returns:
            Path to generated Excel file
        """
        # Create workbook
        self.create_workbook(title)
        
        # Add title
        self.add_title(title, row=1)
        
        # Add headers
        headers = ['Section', 'Content', 'Details']
        self.add_header_row(headers, row=3)
        
        # Add data rows
        current_row = 4
        for i, chunk in enumerate(text_chunks, 1):
            lines = chunk.strip().split('\n')
            
            if lines:
                section_title = lines[0][:100] if len(lines[0]) > 100 else lines[0]
                content = '\n'.join(lines[1:]) if len(lines) > 1 else chunk
                
                # Extract key details if present
                details = self._extract_details(chunk)
                
                data = [f"Section {i}", section_title, content]
                self.add_data_row(data, current_row)
                current_row += 1
        
        # Auto-adjust columns
        self.auto_adjust_column_width()
        
        # Save workbook
        if output_path is None:
            output_dir = os.path.join('output', 'excel')
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, f"{title.replace(' ', '_')}.xlsx")
        
        self.wb.save(output_path)
        return output_path
    
    def generate_structured_data(self, data: List[Dict], title: str,
                                output_path: str = None) -> str:
        """
        Generate Excel from structured data
        
        Args:
            data: List of dictionaries with data
            title: Workbook title
            output_path: Output file path
            
        Returns:
            Path to generated Excel file
        """
        if not data:
            return None
        
        # Create workbook
        self.create_workbook(title)
        
        # Add title
        self.add_title(title, row=1)
        
        # Extract headers from first item
        headers = list(data[0].keys())
        self.add_header_row(headers, row=3)
        
        # Add data rows
        current_row = 4
        for item in data:
            row_data = [str(item.get(key, '')) for key in headers]
            self.add_data_row(row_data, current_row)
            current_row += 1
        
        # Auto-adjust columns
        self.auto_adjust_column_width()
        
        # Save workbook
        if output_path is None:
            output_dir = os.path.join('output', 'excel')
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, f"{title.replace(' ', '_')}.xlsx")
        
        self.wb.save(output_path)
        return output_path
    
    def _extract_details(self, text: str) -> str:
        """
        Extract key details from text
        
        Args:
            text: Text to analyze
            
        Returns:
            Extracted details
        """
        # Simple extraction - look for numbers, dates, etc.
        details = []
        
        # Extract numbers
        numbers = re.findall(r'\d+(?:\.\d+)?%?', text)
        if numbers:
            details.append(f"Numbers: {', '.join(numbers[:3])}")
        
        # Extract any dates
        dates = re.findall(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', text)
        if dates:
            details.append(f"Dates: {', '.join(dates[:2])}")
        
        return '; '.join(details) if details else 'N/A'
